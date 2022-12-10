from openpyxl import load_workbook
wb = load_workbook("sample.xlsx")
ws = wb.active

# 번호, 영어, 수학
a1 = ws["A1"] # 번호
b2 = ws["B1"] # 영어
c1 = ws["C1"] # 수학

ws.column_dimensions["A"].width = 5

wb.save("sample_style.xlsx")