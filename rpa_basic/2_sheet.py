from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성
ws.title = "MySheet" # sheet 이름이 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB형태로 값을 넣어주면 탭 색상 변경

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번쨰 index에 sheet 생성

new_ws = wb["NewSheet"]  # Dict(딕셔너리) 형태로 sheet 에 접근 가능

print(wb.sheetnames) # 현재 열려있는 파일에 있는 모든 sheet이름 출력

# sheet 복사
new_ws["A1"] = "Test"
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet"

wb.save("sample.xlsx")